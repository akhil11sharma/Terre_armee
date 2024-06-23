import os
import openpyxl
import csv
import tkinter as tk
from tkinter import ttk, filedialog
from openpyxl.styles import PatternFill

def display_excel(file_path):
    # Load the Excel workbook and get the active sheet
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active

    # Extract headers from the first row and ensure they are unique
    headers = []
    seen_headers = set()
    for cell in sheet[1]:
        header = cell.value.strip() if cell.value else ""
        if header:
            original_header = header
            count = 1
            while header in seen_headers:
                header = f"{original_header}_{count}"
                count += 1
            headers.append(header)
            seen_headers.add(header)
        else:
            headers.append("")

    # Extract data from all rows, handling different data types
    data = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        row_data = []
        for cell in row:
            if cell.value is None:
                row_data.append("")
            elif isinstance(cell.value, (int, float)):
                row_data.append(str(cell.value))  # Convert numeric values to string
            elif isinstance(cell.value, str):
                row_data.append(cell.value.strip())  # Remove leading/trailing whitespace
            else:
                row_data.append(str(cell.value))  # Convert other types to string
        data.append(row_data)

    workbook.close()

    # Create a tkinter window for displaying the Excel data
    root = tk.Tk()
    root.title("Excel Viewer")

    frame = tk.Frame(root)
    frame.pack(fill=tk.BOTH, expand=True)

    # Create a Treeview widget with extracted headers
    tree = ttk.Treeview(frame, columns=headers, show="headings")

    # Add vertical and horizontal scrollbars
    v_scroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=v_scroll.set)
    h_scroll = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
    tree.configure(xscrollcommand=h_scroll.set)

    # Set up the column headers and properties
    for header in headers:
        if header:
            tree.heading(header, text=header.replace("_", " "))  # Replace underscores with spaces
            tree.column(header, width=100, anchor='center')  # Set column width and alignment

    # Insert the extracted data into the Treeview
    for row in data:
        if all(value == "" for value in row):
            continue  # Skip empty rows
        tree.insert("", "end", values=row)

    # Pack the Treeview and scrollbars into the window
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    h_scroll.pack(side=tk.BOTTOM, fill=tk.X)

    root.mainloop()

def get_formula_columns(file_path):
    # Load the Excel workbook and get the active sheet
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    sheet = workbook.active

    headers = []
    formula_columns = []
    formula_columns_indices = []

    # Identify headers dynamically by scanning the first few rows
    max_header_scan_rows = 5
    for col in sheet.iter_cols(1, sheet.max_column, 1, max_header_scan_rows):
        for cell in col:
            if cell.value:
                headers.append(cell.value)
                break  # Stop scanning once a non-empty header is found

    # Identify columns containing formulas
    for col_idx, col in enumerate(sheet.iter_cols(1, sheet.max_column)):
        for cell in col:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                if col_idx < len(headers):
                    formula_columns.append(headers[col_idx])
                    formula_columns_indices.append(col_idx)
                    break
                else:
                    print(f"Column index {col_idx} is out of range for headers list.")

    workbook.close()
    return formula_columns, formula_columns_indices

def check_and_highlight_non_formula_cells(file_path):
    # Load the Excel workbook and get the active sheet
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    sheet = workbook.active

    headers = []

    # Identify headers dynamically by scanning the first few rows
    max_header_scan_rows = 5
    for col in sheet.iter_cols(1, sheet.max_column, 1, max_header_scan_rows):
        for cell in col:
            if cell.value:
                headers.append(cell.value)
                break  # Stop scanning once a non-empty header is found

    # Get columns containing formulas and their indices
    formula_columns, formula_columns_indices = get_formula_columns(file_path)

    all_columns = list(sheet.iter_cols(1, sheet.max_column))

    # Define the fill style for highlighting cells
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow highlight

    # Highlight non-formula cells in formula columns
    for col_idx in formula_columns_indices:
        for row_idx, cell in enumerate(all_columns[col_idx], start=1):
            if row_idx == 1:
                continue  # Skip the header row
            if cell.value and not (isinstance(cell.value, str) and cell.value.startswith('=')):
                cell_address = openpyxl.utils.get_column_letter(col_idx + 1) + str(row_idx)
                print(f"Cell {cell_address} doesn't have a formula in column {headers[col_idx]}")
                cell.fill = highlight_fill

    workbook.save(file_path)
    workbook.close()

    # Open the Excel file to show the highlighted cells
    os.startfile(file_path)

if __name__ == "__main__":
    # Replace with your file path
    file_path = r"C:\Users\Nikhil Sharma\Desktop\SSMS\areaCalculationSheet.xlsx"
    
    # Display the Excel file in a Tkinter window
    display_excel(file_path)
    
    # Check and highlight non-formula cells in the Excel file
    check_and_highlight_non_formula_cells(file_path)
