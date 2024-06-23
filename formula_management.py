import openpyxl
import pyodbc
import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter import Canvas, Scrollbar

def extract_formulas(file_path):
    # Load the workbook and get the active sheet
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    sheet = workbook.active

    formulas = []
    # Extract the headers from the first row
    headers = [cell.value for cell in next(sheet.iter_rows())]

    # Iterate over the rows to find cells containing formulas
    for row_index, row in enumerate(sheet.iter_rows(min_row=2)):
        for col_index, cell in enumerate(row):
            if cell.data_type == 'f':  # Check if the cell contains a formula
                formula_info = {
                    'Formula': cell.value,
                    'Header': headers[col_index],
                    'CellName': cell.coordinate,
                    'InvolvedColumns': replace_with_headers(cell.value, headers)
                }
                formulas.append(formula_info)

    workbook.close()
    return formulas

def replace_with_headers(formula, headers):
    # Find all column references in the formula
    column_references = re.findall(r'([A-Za-z]+)(\d+)', formula)
    replaced_formula = formula

    # Replace column references with header names
    for col_ref, row_num in column_references:
        column_name = col_ref.upper()
        col_index = openpyxl.utils.cell.column_index_from_string(column_name) - 1

        if col_index < len(headers) and headers[col_index]:
            header_name = headers[col_index]
            replaced_formula = replaced_formula.replace(f"{col_ref}{row_num}", header_name)

    return replaced_formula

def table_exists(cursor, table_name):
    # Check if the table exists in the database
    cursor.execute(f'''
        SELECT COUNT(*)
        FROM information_schema.tables
        WHERE table_name = '{table_name}'
    ''')
    return cursor.fetchone()[0] == 1

def create_tracking_table_if_not_exists(cursor):
    # Create the tracking table if it does not exist
    cursor.execute(f'''
        IF NOT EXISTS (SELECT * FROM sys.tables WHERE name='areaCalculationSheet_Formulas')
        BEGIN
            CREATE TABLE areaCalculationSheet_Formulas (
                Formula NVARCHAR(MAX),
                Column_Header NVARCHAR(255),
                CellName NVARCHAR(255),
                Involved_Columns NVARCHAR(MAX),
                Table_Name NVARCHAR(255)
            );
            PRINT 'Tracking table created successfully: areaCalculationSheet_Formulas';
        END
    ''')

def update_existing_table(cursor, table_name, formulas):
    # Update the existing table with formula data or insert new records
    for formula_info in formulas:
        formula = formula_info['Formula']
        header = formula_info['Header']
        cell_name = formula_info['CellName']
        involved_columns = formula_info['InvolvedColumns']

        cursor.execute(f'''
            IF EXISTS (SELECT * FROM areaCalculationSheet_Formulas WHERE CellName = ? AND Table_Name = ?)
            BEGIN
                UPDATE areaCalculationSheet_Formulas
                SET Formula = ?, Column_Header = ?, Involved_Columns = ?
                WHERE CellName = ? AND Table_Name = ?
            END
            ELSE
            BEGIN
                INSERT INTO areaCalculationSheet_Formulas (Formula, Column_Header, CellName, Involved_Columns, Table_Name)
                VALUES (?, ?, ?, ?, ?)
            END
        ''', (cell_name, table_name, formula, header, involved_columns, cell_name, table_name, formula, header, cell_name, involved_columns, table_name))

def store_in_database(formulas, table_name, db_config):
    # Connect to the database using the provided configuration
    conn_str = f"DRIVER={{SQL Server}};SERVER={db_config['server']};DATABASE={db_config['database']};trusted_connection={db_config['trusted_connection']}"
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    # Ensure the tracking table exists
    create_tracking_table_if_not_exists(cursor)

    # Store the extracted formulas in the database
    update_existing_table(cursor, table_name, formulas)

    cursor.commit()
    conn.close()

def process_file(file_path, db_config):
    # Extract the file name without extension for table name
    file_name = os.path.splitext(os.path.basename(file_path))[0]
    formulas = extract_formulas(file_path)

    if formulas:
        store_in_database(formulas, file_name, db_config)
        return f"File '{file_path}' successfully processed and data stored in table 'areaCalculationSheet_Formulas' under '{file_name}'."
    else:
        return "No formulas found in the Excel sheet."

def select_file():
    # Open a file dialog to select the Excel file
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if file_path:
        selected_file_label.config(text=f"Selected File: {file_path}")
        execute_button.pack()

def execute_process():
    # Execute the process of extracting and storing formulas
    file_path = selected_file_label.cget("text").split(": ")[1]  # Get the selected file path
    message = process_file(file_path, db_config)
    messagebox.showinfo("Process Completed", message)

def display_initial_screen():
    # Display the initial screen of the GUI
    for widget in frame.winfo_children():
        widget.destroy()

    select_label = ttk.Label(frame, text="SELECT THE FILE TO SEND", font=("Helvetica", 16, "bold"))
    select_label.pack(pady=10)

    select_button = ttk.Button(frame, text="Select Excel File", style="Accent.TButton", command=select_file)
    select_button.pack(pady=10)

    global selected_file_label
    selected_file_label = ttk.Label(frame, text="Selected File: None")
    selected_file_label.pack(pady=10)

    global execute_button
    execute_button = ttk.Button(frame, text="Execute", style="Accent.TButton", command=execute_process)

if __name__ == "__main__":
    # Database configuration for connecting to SQL Server
    db_config = {
        'server': 'DESKTOP-PLHJ2M7',
        'database': 'test',
        'trusted_connection': 'yes'
    }

    # Setting up the main Tkinter window
    root = tk.Tk()
    root.title("Formula Extraction")
    root.geometry("600x300")

    # Setting up the main frame and styling
    frame = ttk.Frame(root, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)

    style = ttk.Style()
    style.configure("Accent.TButton", foreground="black", background="orange", font=("Helvetica", 12))

    # Display the initial screen
    display_initial_screen()

    # Start the Tkinter event loop
    root.mainloop()
